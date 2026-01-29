#!/usr/bin/env python3
"""
Script para extrair precos do arquivo Excel e gerar arquivos TypeScript.
Baseado na planilha "monte-sua-casa-simulacao.xlsx"

Uso: python3 scripts/extract-excel.py

Este script le o arquivo Excel e gera:
- src/lib/prices/orcamento-casa.ts (materiais da casa)
- src/lib/prices/mao-obra-casa.ts (mao de obra da casa)
- src/lib/prices/types.ts (interfaces comuns)
"""

import os
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERRO: openpyxl nao encontrado. Instale com: pip install openpyxl")
    sys.exit(1)

# Caminhos
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
EXCEL_FILE = PROJECT_ROOT / "monte-sua-casa-simulacao.xlsx"
PRICES_DIR = SCRIPT_DIR.parent / "src" / "lib" / "prices"

# Constantes da planilha
FATOR_AJUSTE_MATERIAIS = 0.0079  # 0.79%
BDI_PERCENTUAL = 14.40

def safe_float(value, default=0.0):
    """Converte valor para float de forma segura"""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

def safe_str(value, default=""):
    """Converte valor para string de forma segura"""
    if value is None:
        return default
    return str(value).strip()

def extract_orcamento_casa(ws):
    """Extrai dados da aba ORCAMENTO - CASA"""

    # Estrutura para armazenar os precos
    precos = {
        "movimentoTerra": {},
        "baldrameAlvenaria": {},
        "fundacoesEstruturas": {},
        "esquadriasFerragens": {},
        "cobertura": {},
        "revestimentos": {
            "parede": {},
            "teto": {},
            "pisos": {}
        },
        "instalacaoHidraulica": {},
        "instalacaoSanitaria": {},
        "instalacaoEletrica": {},
        "gasGlp": {},
        "pintura": {},
        "churrasqueira": {},
        "limpezaObra": {}
    }

    # Mapeamento de codigos para campos TypeScript
    mapeamento = {
        # 3.1 MOVIMENTO DE TERRA
        "3.1.1": ("movimentoTerra", "escavacaoValasBaldrame", None),
        "3.1.2": ("movimentoTerra", "escavacaoFundacao60x60", None),
        "3.1.3": ("movimentoTerra", "reterroCompactacao", None),
        "3.1.4": ("movimentoTerra", "espalhamentoBase", None),
        "3.1.5": ("movimentoTerra", "apiloamentoFundoVala", None),

        # 3.2 BALDRAME E ALVENARIA
        "3.2.1": ("baldrameAlvenaria", "alvenariaPedraArgamassada", None),
        "3.2.2": ("baldrameAlvenaria", "cintaConcretoArmado", None),
        "3.2.3": ("baldrameAlvenaria", "impermeabilizacaoBaldrame", None),
        "3.2.4": ("baldrameAlvenaria", "alvenariaTijoloFurado", None),

        # 3.3 FUNDACOES E ESTRUTURAS
        "3.3.1": ("fundacoesEstruturas", "concretoPilaresVigas", None),
        "3.3.2": ("fundacoesEstruturas", "formaDesforma", None),
        "3.3.3": ("fundacoesEstruturas", "armaduraCA50", None),
        "3.3.4": ("fundacoesEstruturas", "lancamentoConcreto", None),
        "3.3.5": ("fundacoesEstruturas", "lajePrefabricada", None),

        # 3.4 ESQUADRIAS E FERRAGENS
        "3.4.1": ("esquadriasFerragens", "portaEntradaDecorativa", None),
        "3.4.2": ("esquadriasFerragens", "portaMadeiraLei", None),
        "3.4.3": ("esquadriasFerragens", "janelaAluminio", None),  # Aluminio e vidro
        "3.4.4": ("esquadriasFerragens", "cobogoAntiChuva", None),

        # 3.5 COBERTURA
        "3.5.1": ("cobertura", "cobertaPadrao", None),

        # 3.6.1 REVESTIMENTOS - PAREDE
        "3.6.1.1": ("revestimentos", "chapiscoCimentoAreia", "parede"),
        "3.6.1.2": ("revestimentos", "rebocoCimentoAreia", "parede"),
        "3.6.1.3": ("revestimentos", "embocoCimentoAreia", "parede"),
        "3.6.1.4": ("revestimentos", "revestimentoCeramico", "parede"),
        "3.6.1.13": ("revestimentos", "rejuntamentoPorcelanato", "parede"),
        "3.6.1.14": ("revestimentos", "bancadaCozinhaPorcelanato", "parede"),

        # 3.6.2 REVESTIMENTOS - TETO
        "3.6.2.1": ("revestimentos", "gessoConvencionalForro", "teto"),

        # 3.6.3 REVESTIMENTOS - PISOS
        "3.6.3.1": ("revestimentos", "concretoNaoEstruturalLastro", "pisos"),
        "3.6.3.2": ("revestimentos", "regularizacaoBase", "pisos"),
        "3.6.3.3": ("revestimentos", "revestimentoCeramico", "pisos"),
        "3.6.3.28": ("revestimentos", "rejuntamentoPorcelanato", "pisos"),
        "3.6.3.29": ("revestimentos", "soleirasGranito", "pisos"),

        # 3.7 INSTALACAO HIDRAULICA
        "3.7.1": ("instalacaoHidraulica", "tuboPVC50mm", None),
        "3.7.2": ("instalacaoHidraulica", "tuboPVC32mm", None),
        "3.7.3": ("instalacaoHidraulica", "tuboPVC25mm", None),
        "3.7.4": ("instalacaoHidraulica", "caixaDagua1500L", None),
        "3.7.5": ("instalacaoHidraulica", "flange2pol", None),
        "3.7.6": ("instalacaoHidraulica", "flange1pol", None),
        "3.7.7": ("instalacaoHidraulica", "registroGaveta", None),
        "3.7.8": ("instalacaoHidraulica", "registroGavetaCanopla", None),
        "3.7.9": ("instalacaoHidraulica", "registroPressaoChuveiro", None),
        "3.7.10": ("instalacaoHidraulica", "boiaMecanica", None),
        "3.7.11": ("instalacaoHidraulica", "torneiraMetal", None),
        "3.7.12": ("instalacaoHidraulica", "bancadaGranitoLavatorio", None),
        "3.7.13": ("instalacaoHidraulica", "baciaSanitaria", None),
        "3.7.14": ("instalacaoHidraulica", "chuveiroArticulado", None),
        "3.7.15": ("instalacaoHidraulica", "bancadaGranitoCozinha", None),
        "3.7.17": ("instalacaoHidraulica", "tanqueInox", None),

        # 3.8 INSTALACAO SANITARIA
        "3.8.1": ("instalacaoSanitaria", "caixaInspecao60x60", None),
        "3.8.2": ("instalacaoSanitaria", "tuboPVCEsgoto100mm", None),
        "3.8.3": ("instalacaoSanitaria", "tuboPVCEsgoto75mm", None),
        "3.8.5": ("instalacaoSanitaria", "tuboPVCEsgoto50mm", None),
        "3.8.6": ("instalacaoSanitaria", "raloSifonado", None),

        # 3.9 INSTALACAO ELETRICA
        "3.9.1": ("instalacaoEletrica", "quadroDistribuicao12", None),
        "3.9.2": ("instalacaoEletrica", "eletrodutoRigido32mm", None),
        "3.9.3": ("instalacaoEletrica", "eletrodutoFlexivel", None),
        "3.9.4": ("instalacaoEletrica", "caixaLigacaoPVC4x4", None),
        "3.9.5": ("instalacaoEletrica", "caixaLigacaoPVC4x2", None),
        "3.9.6": ("instalacaoEletrica", "caboIsoladoPVC1_5mm", None),
        "3.9.7": ("instalacaoEletrica", "caboIsoladoPVC2_5mm", None),
        "3.9.8": ("instalacaoEletrica", "caboIsoladoPVC4mm", None),
        "3.9.9": ("instalacaoEletrica", "caboIsoladoPVC10mm", None),
        "3.9.10": ("instalacaoEletrica", "disjuntor15A", None),
        "3.9.11": ("instalacaoEletrica", "disjuntor20A", None),
        "3.9.12": ("instalacaoEletrica", "disjuntor32A", None),
        "3.9.13": ("instalacaoEletrica", "disjuntor50A", None),
        "3.9.14": ("instalacaoEletrica", "hasteCobre", None),
        "3.9.15": ("instalacaoEletrica", "interruptorTriplo", None),
        "3.9.16": ("instalacaoEletrica", "interruptorDuplo", None),
        "3.9.18": ("instalacaoEletrica", "interruptorCampainha", None),
        "3.9.19": ("instalacaoEletrica", "tomadaTripla", None),
        "3.9.20": ("instalacaoEletrica", "pontoLogica", None),
        "3.9.21": ("instalacaoEletrica", "pontoTV", None),
        "3.9.22": ("instalacaoEletrica", "luminariaLED", None),

        # 3.10 GAS GLP
        "3.10.1": ("gasGlp", "tuboCobre15mm", None),
        "3.10.2": ("gasGlp", "testeEstanqueidade", None),

        # 3.11 PINTURA
        "3.11.2": ("pintura", "texturaExterna", None),
        "3.11.4": ("pintura", "emassamento", None),
        "3.11.5": ("pintura", "pinturaLatexPVA", None),
        "3.11.6": ("pintura", "seladorMadeira", None),
        "3.11.7": ("pintura", "esmalteSintetico", None),

        # 3.12 CHURRASQUEIRA
        "3.12.1": ("churrasqueira", "churrasqueiraMediaPorte", None),

        # 3.13 LIMPEZA DA OBRA
        "3.13.1": ("limpezaObra", "containers", None),
        "3.13.2": ("limpezaObra", "transporteHorizontal", None),
        "3.13.3": ("limpezaObra", "limpezaGeral", None),
    }

    # Percorre as linhas da planilha
    for row in ws.iter_rows(min_row=7, max_row=150):
        codigo = safe_str(row[1].value)  # Coluna B (indice 1)
        preco_base = safe_float(row[7].value)  # Coluna H (indice 7)

        if codigo in mapeamento and preco_base > 0:
            secao, campo, subsecao = mapeamento[codigo]

            if subsecao:
                if secao == "revestimentos":
                    precos[secao][subsecao][campo] = preco_base
            else:
                precos[secao][campo] = preco_base

    return precos

def extract_mao_obra_casa(ws):
    """Extrai dados da aba MAO DE OBRA - CASA

    Estrutura desta aba e diferente: usa descricao direta na coluna B
    em vez de codigos numericos como 3.1.1
    """

    precos = {
        "movimentoTerra": {},
        "baldrameAlvenaria": {},
        "fundacoesEstruturas": {},
        "esquadriasFerragens": {},
        "cobertura": {},
        "revestimentos": {
            "parede": {},
            "teto": {},
            "pisos": {}
        },
        "instalacaoHidraulica": {},
        "instalacaoSanitaria": {},
        "instalacaoEletrica": {},
        "gasGlp": {},
        "pintura": {},
        "churrasqueira": {},
        "limpezaObra": {}
    }

    # Mapeamento de descricoes para campos TypeScript
    # Normalizado para lowercase e sem acentos para comparacao
    mapeamento_mo = {
        # 3.1 MOVIMENTO DE TERRA
        "escavacao manual de valas": ("movimentoTerra", "escavacaoValasBaldrame", None),
        "escavacao manual de fundacao": ("movimentoTerra", "escavacaoFundacao60x60", None),
        "reterro manual": ("movimentoTerra", "reterroCompactacao", None),
        "espalhamento e adensamento": ("movimentoTerra", "espalhamentoBase", None),
        "apiloamento de fundo": ("movimentoTerra", "apiloamentoFundoVala", None),

        # 3.2 BALDRAME E ALVENARIA
        "alvenaria de pedra argamassada": ("baldrameAlvenaria", "alvenariaPedraArgamassada", None),
        "cinta em concreto armado": ("baldrameAlvenaria", "cintaConcretoArmado", None),
        "impermeabilizacao de baldrame": ("baldrameAlvenaria", "impermeabilizacaoBaldrame", None),
        "alvenaria em tijolo furado": ("baldrameAlvenaria", "alvenariaTijoloFurado", None),

        # 3.3 FUNDACOES E ESTRUTURAS
        "concreto em pilares": ("fundacoesEstruturas", "concretoPilaresVigas", None),
        "forma e desforma": ("fundacoesEstruturas", "formaDesforma", None),
        "armadura ca 50": ("fundacoesEstruturas", "armaduraCA50", None),
        "lancamento e aplicacao": ("fundacoesEstruturas", "lancamentoConcreto", None),
        "lajes pre-fabricada": ("fundacoesEstruturas", "lajePrefabricada", None),

        # 3.4 ESQUADRIAS E FERRAGENS
        "porta de entrada decorativa": ("esquadriasFerragens", "portaEntradaDecorativa", None),
        "porta de madeira de lei": ("esquadriasFerragens", "portaMadeiraLei", None),
        "esquadria de aluminio": ("esquadriasFerragens", "janelaAluminio", None),
        "cobogo": ("esquadriasFerragens", "cobogoAntiChuva", None),

        # 3.5 COBERTURA
        "coberta de acordo com briefing": ("cobertura", "cobertaPadrao", None),

        # 3.6.1 REVESTIMENTOS - PAREDE
        "chapisco traco cimento e areia": ("revestimentos", "chapiscoCimentoAreia", "parede"),
        "reboco cimento e areia": ("revestimentos", "rebocoCimentoAreia", "parede"),
        "emboco cimento e areia": ("revestimentos", "embocoCimentoAreia", "parede"),
        "bancada da cozinha": ("revestimentos", "bancadaCozinhaPorcelanato", "parede"),

        # 3.6.2 REVESTIMENTOS - TETO
        "gesso convencional para forro": ("revestimentos", "gessoConvencionalForro", "teto"),

        # 3.6.3 REVESTIMENTOS - PISOS
        "concreto nao estrutural": ("revestimentos", "concretoNaoEstruturalLastro", "pisos"),
        "regularizacao de base": ("revestimentos", "regularizacaoBase", "pisos"),
        "soleiras de granito": ("revestimentos", "soleirasGranito", "pisos"),

        # 3.7 INSTALACAO HIDRAULICA
        "tubo soldavel em pvc 25mm": ("instalacaoHidraulica", "tuboPVC25mm", None),
        "tubo soldavel em pvc 32mm": ("instalacaoHidraulica", "tuboPVC32mm", None),
        "tubo soldavel em pvc 50mm": ("instalacaoHidraulica", "tuboPVC50mm", None),
        "caixa d'agua": ("instalacaoHidraulica", "caixaDagua1500L", None),
        "flange 2": ("instalacaoHidraulica", "flange2pol", None),
        "flange de 1": ("instalacaoHidraulica", "flange1pol", None),
        "registro bruto de gaveta": ("instalacaoHidraulica", "registroGaveta", None),
        "registro de gaveta c/ canopla": ("instalacaoHidraulica", "registroGavetaCanopla", None),
        "registro de presao para chuveiro": ("instalacaoHidraulica", "registroPressaoChuveiro", None),
        "boia mecanica": ("instalacaoHidraulica", "boiaMecanica", None),
        "torneira para jardim": ("instalacaoHidraulica", "torneiraMetal", None),
        "bancada de granito para lavatorio": ("instalacaoHidraulica", "bancadaGranitoLavatorio", None),
        "bacia sanitaria": ("instalacaoHidraulica", "baciaSanitaria", None),
        "ducha higienica": ("instalacaoHidraulica", "duchaHigienica", None),
        "chuveiro articulado": ("instalacaoHidraulica", "chuveiroArticulado", None),
        "bancada em granito p/ pia de cozinha": ("instalacaoHidraulica", "bancadaGranitoCozinha", None),
        "tanque de inox": ("instalacaoHidraulica", "tanqueInox", None),

        # 3.8 INSTALACAO SANITARIA
        "caixa de inspecao em alvenaria": ("instalacaoSanitaria", "caixaInspecao60x60", None),
        "tubo e conexao em pvc para esgoto 100mm": ("instalacaoSanitaria", "tuboPVCEsgoto100mm", None),
        "tubo e conexao em pvc para esgoto 75mm": ("instalacaoSanitaria", "tuboPVCEsgoto75mm", None),
        "tubo e conexao em pvc para esgoto 50mm": ("instalacaoSanitaria", "tuboPVCEsgoto50mm", None),
        "ralo sofonado": ("instalacaoSanitaria", "raloSifonado", None),

        # 3.9 INSTALACAO ELETRICA
        "quadro de distribuicao": ("instalacaoEletrica", "quadroDistribuicao12", None),
        "eletroduto rigido": ("instalacaoEletrica", "eletrodutoRigido32mm", None),
        "eletroduto flexivel": ("instalacaoEletrica", "eletrodutoFlexivel", None),
        "caixa de ligacao em pvc rigido 4x4": ("instalacaoEletrica", "caixaLigacaoPVC4x4", None),
        "caixa de ligacao em pvc rigido 4x2": ("instalacaoEletrica", "caixaLigacaoPVC4x2", None),
        "cabo isolado pvc, 750 v, - 1,5mm": ("instalacaoEletrica", "caboIsoladoPVC1_5mm", None),
        "cabo isolado pvc, 750 v, - 2,5mm": ("instalacaoEletrica", "caboIsoladoPVC2_5mm", None),
        "cabo isolado pvc, 750 v, - 4,0mm": ("instalacaoEletrica", "caboIsoladoPVC4mm", None),
        "cabo isolado pvc, 750 v, - 10": ("instalacaoEletrica", "caboIsoladoPVC10mm", None),
        "disjuntores 15a": ("instalacaoEletrica", "disjuntor15A", None),
        "disjuntor 20a": ("instalacaoEletrica", "disjuntor20A", None),
        "disjuntores 32a": ("instalacaoEletrica", "disjuntor32A", None),
        "disjuntor de 50a": ("instalacaoEletrica", "disjuntor50A", None),
        "haste de cobre": ("instalacaoEletrica", "hasteCobre", None),
        "interruptor triplo": ("instalacaoEletrica", "interruptorTriplo", None),
        "interruptor duplo": ("instalacaoEletrica", "interruptorDuplo", None),
        "interruptos para capainha": ("instalacaoEletrica", "interruptorCampainha", None),
        "tomada tripla": ("instalacaoEletrica", "tomadaTripla", None),
        "ponto de logica": ("instalacaoEletrica", "pontoLogica", None),
        "ponto de televisao": ("instalacaoEletrica", "pontoTV", None),
        "luminaria de led": ("instalacaoEletrica", "luminariaLED", None),

        # 3.10 GAS GLP
        "tubo de cobre d=15mm": ("gasGlp", "tuboCobre15mm", None),
        "teste de estanqueidade": ("gasGlp", "testeEstanqueidade", None),

        # 3.11 PINTURA
        "textura duas demaos externa": ("pintura", "texturaExterna", None),
        "emassamento duas demaos": ("pintura", "emassamento", None),
        "latex interno duas demaos": ("pintura", "pinturaLatexPVA", None),
        "selador em madeira": ("pintura", "seladorMadeira", None),
        "esmalte sintetico duas demaos": ("pintura", "esmalteSintetico", None),

        # 3.12 CHURRASQUEIRA
        "churrasqueira medio porte": ("churrasqueira", "churrasqueiraMediaPorte", None),

        # 3.13 LIMPEZA DA OBRA
        "transporte horizontal": ("limpezaObra", "transporteHorizontal", None),
        "limpeza geral": ("limpezaObra", "limpezaGeral", None),
    }

    # Seção atual para tratar revestimentos ceramico e rejuntamento
    secao_revestimento_atual = None

    # Percorre as linhas
    for row in ws.iter_rows(min_row=7, max_row=120):
        descricao_original = safe_str(row[1].value)  # Coluna B
        preco = safe_float(row[7].value)  # Coluna H (preco unitario)

        if not descricao_original:
            continue

        # Normaliza descricao para comparacao
        descricao = descricao_original.lower()
        descricao = descricao.replace('ã', 'a').replace('á', 'a').replace('â', 'a')
        descricao = descricao.replace('é', 'e').replace('ê', 'e')
        descricao = descricao.replace('í', 'i')
        descricao = descricao.replace('ó', 'o').replace('ô', 'o')
        descricao = descricao.replace('ú', 'u')
        descricao = descricao.replace('ç', 'c')

        # Detecta secao de revestimento
        if descricao == "parede":
            secao_revestimento_atual = "parede"
            continue
        elif descricao == "teto":
            secao_revestimento_atual = "teto"
            continue
        elif descricao == "pisos":
            secao_revestimento_atual = "pisos"
            continue

        # Trata revestimento ceramico e rejuntamento baseado na secao atual
        if "revestimento ceramico" in descricao and secao_revestimento_atual and preco > 0:
            precos["revestimentos"][secao_revestimento_atual]["revestimentoCeramico"] = preco
            continue
        if "rejuntamento para porcelanato" in descricao and secao_revestimento_atual and preco > 0:
            precos["revestimentos"][secao_revestimento_atual]["rejuntamentoPorcelanato"] = preco
            continue

        # Procura correspondencia no mapeamento
        if preco > 0:
            for chave, valor in mapeamento_mo.items():
                if chave in descricao:
                    secao, campo, subsecao = valor
                    if subsecao:
                        precos[secao][subsecao][campo] = preco
                    else:
                        precos[secao][campo] = preco
                    break

    return precos

def generate_types_ts():
    """Gera o arquivo types.ts com interfaces comuns"""
    return f'''// Tipos e interfaces para precos - Gerado automaticamente
// Data de geracao: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
// Fonte: monte-sua-casa-simulacao.xlsx

/**
 * Fator de ajuste para materiais (0.79%)
 * Aplicado sobre o preco base para obter o preco ajustado
 */
export const FATOR_AJUSTE_MATERIAIS = {FATOR_AJUSTE_MATERIAIS};

/**
 * Aplica o fator de ajuste ao preco base
 * @param precoBase - Preco base do item
 * @returns Preco ajustado com fator de 0.79%
 */
export function aplicarAjuste(precoBase: number): number {{
  return precoBase * (1 + FATOR_AJUSTE_MATERIAIS);
}}

/**
 * Interface para precos de uma secao
 */
export interface PrecosSecao {{
  [key: string]: number;
}}

/**
 * Interface para sub-secoes de revestimentos
 */
export interface PrecosRevestimentos {{
  parede: PrecosSecao;
  teto: PrecosSecao;
  pisos: PrecosSecao;
}}

/**
 * Interface completa de precos de materiais
 */
export interface PrecosMateriais {{
  movimentoTerra: PrecosSecao;
  baldrameAlvenaria: PrecosSecao;
  fundacoesEstruturas: PrecosSecao;
  esquadriasFerragens: PrecosSecao;
  cobertura: PrecosSecao;
  revestimentos: PrecosRevestimentos;
  instalacaoHidraulica: PrecosSecao;
  instalacaoSanitaria: PrecosSecao;
  instalacaoEletrica: PrecosSecao;
  gasGlp: PrecosSecao;
  pintura: PrecosSecao;
  churrasqueira: PrecosSecao;
  limpezaObra: PrecosSecao;
}}

/**
 * Interface completa de precos de mao de obra
 */
export interface PrecosMaoObra {{
  movimentoTerra: PrecosSecao;
  baldrameAlvenaria: PrecosSecao;
  fundacoesEstruturas: PrecosSecao;
  esquadriasFerragens: PrecosSecao;
  cobertura: PrecosSecao;
  revestimentos: PrecosRevestimentos;
  instalacaoHidraulica: PrecosSecao;
  instalacaoSanitaria: PrecosSecao;
  instalacaoEletrica: PrecosSecao;
  gasGlp: PrecosSecao;
  pintura: PrecosSecao;
  churrasqueira: PrecosSecao;
  limpezaObra: PrecosSecao;
}}
'''

def generate_orcamento_casa_ts(precos):
    """Gera o arquivo orcamento-casa.ts com os precos extraidos"""

    sections = []
    for section_name, section_data in precos.items():
        if isinstance(section_data, dict) and any(isinstance(v, dict) for v in section_data.values()):
            # Secao com sub-secoes (revestimentos)
            subsections = []
            for sub_name, sub_data in section_data.items():
                sub_items = [f"      {k}: {v}," for k, v in sub_data.items() if v > 0]
                if sub_items:
                    subsections.append(f"    {sub_name}: {{\n" + "\n".join(sub_items) + "\n    },")
                else:
                    subsections.append(f"    {sub_name}: {{}},")
            sections.append(f"  {section_name}: {{\n" + "\n".join(subsections) + "\n  },")
        else:
            # Secao simples
            items = [f"    {k}: {v}," for k, v in section_data.items() if v > 0]
            if items:
                sections.append(f"  {section_name}: {{\n" + "\n".join(items) + "\n  },")
            else:
                sections.append(f"  {section_name}: {{}},")

    return f'''// Precos de materiais da Casa - Extraido automaticamente do Excel
// Data de geracao: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
// Fonte: monte-sua-casa-simulacao.xlsx - Aba "ORCAMENTO - CASA"
//
// IMPORTANTE: Este arquivo e gerado automaticamente pelo script extract-excel.py
// Nao edite manualmente. Para atualizar, modifique o Excel e execute o script.

import {{ PrecosMateriais }} from './types';

/**
 * Precos base de materiais da casa (sem ajuste)
 * O fator de ajuste (0.79%) deve ser aplicado ao usar estes precos
 */
export const PRECOS_MATERIAIS_CASA: PrecosMateriais = {{
{chr(10).join(sections)}
}};

// Exporta tipos
export * from './types';
'''

def generate_mao_obra_casa_ts(precos):
    """Gera o arquivo mao-obra-casa.ts com os precos extraidos"""

    sections = []
    for section_name, section_data in precos.items():
        if isinstance(section_data, dict) and any(isinstance(v, dict) for v in section_data.values()):
            subsections = []
            for sub_name, sub_data in section_data.items():
                sub_items = [f"      {k}: {v}," for k, v in sub_data.items() if v > 0]
                if sub_items:
                    subsections.append(f"    {sub_name}: {{\n" + "\n".join(sub_items) + "\n    },")
                else:
                    subsections.append(f"    {sub_name}: {{}},")
            sections.append(f"  {section_name}: {{\n" + "\n".join(subsections) + "\n  },")
        else:
            items = [f"    {k}: {v}," for k, v in section_data.items() if v > 0]
            if items:
                sections.append(f"  {section_name}: {{\n" + "\n".join(items) + "\n  },")
            else:
                sections.append(f"  {section_name}: {{}},")

    return f'''// Precos de mao de obra da Casa - Extraido automaticamente do Excel
// Data de geracao: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
// Fonte: monte-sua-casa-simulacao.xlsx - Aba "MAO DE OBRA - CASA"
//
// IMPORTANTE: Este arquivo e gerado automaticamente pelo script extract-excel.py
// Nao edite manualmente. Para atualizar, modifique o Excel e execute o script.

import {{ PrecosMaoObra }} from './types';

/**
 * BDI (Beneficios e Despesas Indiretas) percentual
 * Conforme planilha: 14.40%
 */
export const BDI_PERCENTUAL = {BDI_PERCENTUAL};

/**
 * Precos de mao de obra da casa
 */
export const PRECOS_MAO_OBRA_CASA: PrecosMaoObra = {{
{chr(10).join(sections)}
}};

// Exporta tipos
export * from './types';
'''

def generate_index_ts():
    """Gera o arquivo index.ts que exporta tudo"""
    return f'''// Exportacoes centralizadas de precos
// Data de geracao: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

export * from './types';
export * from './orcamento-casa';
export * from './mao-obra-casa';
'''

def main():
    print("=" * 60)
    print("Extrator de Precos do Excel")
    print("=" * 60)

    # Verifica se arquivo existe
    if not EXCEL_FILE.exists():
        print(f"ERRO: Arquivo Excel nao encontrado: {EXCEL_FILE}")
        sys.exit(1)

    print(f"\nArquivo Excel: {EXCEL_FILE}")
    print(f"Diretorio de saida: {PRICES_DIR}")

    # Cria diretorio se nao existir
    PRICES_DIR.mkdir(parents=True, exist_ok=True)

    # Carrega o workbook
    print("\nCarregando arquivo Excel...")
    wb = load_workbook(EXCEL_FILE, data_only=True)

    print(f"Abas encontradas: {wb.sheetnames}")

    # Extrai dados do ORCAMENTO - CASA
    print("\nExtraindo dados de ORCAMENTO - CASA...")
    if "ORÇAMENTO - CASA" in wb.sheetnames:
        ws_orcamento = wb["ORÇAMENTO - CASA"]
        precos_materiais = extract_orcamento_casa(ws_orcamento)

        # Conta itens extraidos
        count = 0
        for secao, dados in precos_materiais.items():
            if isinstance(dados, dict):
                if any(isinstance(v, dict) for v in dados.values()):
                    for sub in dados.values():
                        count += len([v for v in sub.values() if v > 0])
                else:
                    count += len([v for v in dados.values() if v > 0])
        print(f"  Itens extraidos: {count}")
    else:
        print("  AVISO: Aba 'ORCAMENTO - CASA' nao encontrada")
        precos_materiais = {}

    # Extrai dados da MAO DE OBRA - CASA
    print("\nExtraindo dados de MAO DE OBRA - CASA...")
    if "MÃO DE OBRA - CASA" in wb.sheetnames:
        ws_mao_obra = wb["MÃO DE OBRA - CASA"]
        precos_mao_obra = extract_mao_obra_casa(ws_mao_obra)

        # Conta itens extraidos
        count = 0
        for secao, dados in precos_mao_obra.items():
            if isinstance(dados, dict):
                if any(isinstance(v, dict) for v in dados.values()):
                    for sub in dados.values():
                        count += len([v for v in sub.values() if v > 0])
                else:
                    count += len([v for v in dados.values() if v > 0])
        print(f"  Itens extraidos: {count}")
    else:
        print("  AVISO: Aba 'MAO DE OBRA - CASA' nao encontrada")
        precos_mao_obra = {}

    # Gera arquivos TypeScript
    print("\nGerando arquivos TypeScript...")

    # types.ts
    types_file = PRICES_DIR / "types.ts"
    with open(types_file, "w", encoding="utf-8") as f:
        f.write(generate_types_ts())
    print(f"  Criado: {types_file}")

    # orcamento-casa.ts
    orcamento_file = PRICES_DIR / "orcamento-casa.ts"
    with open(orcamento_file, "w", encoding="utf-8") as f:
        f.write(generate_orcamento_casa_ts(precos_materiais))
    print(f"  Criado: {orcamento_file}")

    # mao-obra-casa.ts
    mao_obra_file = PRICES_DIR / "mao-obra-casa.ts"
    with open(mao_obra_file, "w", encoding="utf-8") as f:
        f.write(generate_mao_obra_casa_ts(precos_mao_obra))
    print(f"  Criado: {mao_obra_file}")

    # index.ts
    index_file = PRICES_DIR / "index.ts"
    with open(index_file, "w", encoding="utf-8") as f:
        f.write(generate_index_ts())
    print(f"  Criado: {index_file}")

    print("\n" + "=" * 60)
    print("Extracao concluida com sucesso!")
    print("=" * 60)

if __name__ == "__main__":
    main()
